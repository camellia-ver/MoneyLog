# -*- coding: utf-8 -*-
"""
MoneyLog QA - 테스트케이스.xlsx 생성 스크립트 (STEP 2)
TC목록 시트 + 요약 시트를 생성한다.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# TC 데이터 정의
# 컬럼: TC-ID | 카테고리 | 테스트 항목 | 전제 조건 | 테스트 절차 | 기대 결과 | 우선순위 | 결과 | 비고
# ---------------------------------------------------------------------------

CRIT = "Critical"
HIGH = "High"
MED = "Medium"
LOW = "Low"

tcs = []

def add(cat, item, pre, steps, expected, priority, note=""):
    tcs.append([None, cat, item, pre, steps, expected, priority, "미실행", note])

# ===== 1. 회원가입 =====
add("회원가입", "정상 회원가입 성공",
    "회원가입 화면 또는 API 접근 가능 상태",
    "1) POST /api/users/signup 호출\n2) email/userName/password(8~20자) 정상값 입력\n3) 요청 전송",
    "201 Created 반환, 응답 바디에 id/email/userName 포함(비밀번호 미노출)",
    HIGH)

add("회원가입", "이메일 중복 가입 시도",
    "동일 이메일로 이미 가입된 계정 존재",
    "1) 기존 가입 이메일로 회원가입 재시도",
    "409 Conflict, '이미 사용 중인 이메일입니다' 메시지 반환",
    HIGH)

add("회원가입", "이메일 형식 오류",
    "-",
    "1) email='invalid-email' 형식으로 회원가입 요청",
    "400 Bad Request, '올바른 이메일 형식이 아닙니다' 메시지 반환",
    MED)

add("회원가입", "이메일 공백/누락",
    "-",
    "1) email 필드를 빈 문자열 또는 미포함으로 회원가입 요청",
    "400 Bad Request, '이메일은 필수입니다' 메시지 반환",
    MED)

add("회원가입", "비밀번호 8자 미만 입력",
    "-",
    "1) password='abc123'(6자)로 회원가입 요청",
    "400 Bad Request, '비밀번호는 8자 이상 20자 이하로 입력해주세요' 메시지 반환",
    MED)

add("회원가입", "비밀번호 20자 초과 입력",
    "-",
    "1) password 21자 이상 문자열로 회원가입 요청",
    "400 Bad Request, 비밀번호 길이 검증 메시지 반환",
    MED)

add("회원가입", "사용자명(userName) 공백",
    "-",
    "1) userName 빈 문자열로 회원가입 요청",
    "400 Bad Request, '사용자명은 필수입니다' 메시지 반환",
    MED)

add("회원가입", "이메일 대소문자 중복 처리 확인",
    "'test@example.com' 계정 존재",
    "1) 'TEST@example.com'(대문자)으로 동일 인물 재가입 시도",
    "정책 확인 필요: 대소문자 무시하고 중복(409) 처리되는지, 별개 계정으로 생성되는지 실제 동작 확인",
    MED, "STEP1 예측 결함: existsByEmail 대소문자 구분 정책 불명확(코드 기반 미확정)")

# ===== 2. 로그인/인증 =====
add("로그인/인증", "정상 로그인 성공 (JWT 발급)",
    "가입된 계정 존재",
    "1) POST /api/users/login 호출, 정상 email/password 입력",
    "200 OK, token/userId/expiresIn 포함 응답 반환",
    CRIT)

add("로그인/인증", "존재하지 않는 이메일로 로그인",
    "-",
    "1) 가입되지 않은 이메일로 로그인 시도",
    "401 Unauthorized, '이메일 또는 비밀번호가 일치하지 않습니다' 메시지 반환",
    HIGH)

add("로그인/인증", "비밀번호 불일치 로그인",
    "가입된 계정 존재",
    "1) 정상 이메일 + 잘못된 비밀번호로 로그인 시도",
    "401 Unauthorized, 동일한 에러 메시지(이메일/비밀번호 구분 노출 금지) 반환",
    HIGH)

add("로그인/인증", "이메일 공백으로 로그인 시도 (DTO 검증 부재 확인)",
    "-",
    "1) email='' , password='정상값'으로 로그인 요청 전송",
    "AuthDto.Request에 Bean Validation이 없어 서비스 레이어까지 그대로 진입할 수 있음. 400(Validation) 또는 401 중 실제 응답과 에러 처리의 일관성을 확인",
    HIGH, "STEP1 예측 결함: AuthDto.Request에 @NotBlank/@Email 없음 → 검증 우회 가능성")

add("로그인/인증", "비밀번호 공백으로 로그인 시도",
    "-",
    "1) email='정상값', password=''로 로그인 요청 전송",
    "적절한 4xx 응답 및 명확한 에러 메시지 반환 여부 확인(500 등 비정상 응답 발생 시 결함)",
    HIGH, "STEP1 예측 결함: AuthDto.Request 검증 부재")

add("로그인/인증", "토큰 없이 보호된 API 호출",
    "-",
    "1) Authorization 헤더 없이 GET /api/categories 등 보호된 API 호출",
    "401 Unauthorized 반환",
    CRIT)

add("로그인/인증", "잘못된 형식의 Authorization 헤더",
    "-",
    "1) Authorization 헤더에 'Bearer ' 접두어 없이 토큰만 전송하거나 'Basic xxx' 형식으로 전송",
    "인증 미적용 상태로 처리되어 401 반환(서버 오류 발생하지 않아야 함)",
    MED)

add("로그인/인증", "위조/변조된 JWT 토큰 사용",
    "정상 발급받은 토큰 보유",
    "1) 정상 토큰의 payload 또는 서명 일부를 임의로 변경\n2) 변경된 토큰으로 보호된 API 호출",
    "validateToken()이 false 반환 → 인증 미적용 → 401 Unauthorized",
    CRIT)

add("로그인/인증", "만료된 JWT 토큰 사용",
    "jwt.expiration을 짧게 설정하거나 만료 대기 가능한 토큰 확보",
    "1) 토큰 만료 시간 경과 후 보호된 API 호출",
    "401 Unauthorized 반환, 만료 관련 처리로 서버 오류(500) 발생하지 않아야 함",
    CRIT)

add("로그인/인증", "로그인 다회 실패 시 계정 잠금/제한 여부",
    "가입된 계정 존재",
    "1) 동일 계정으로 잘못된 비밀번호로 10회 이상 연속 로그인 시도",
    "현재 코드상 별도의 rate limiting/계정 잠금 로직 없음 → 무제한 재시도 가능 여부 확인 및 보안 리스크로 기록",
    MED, "STEP1 리스크: 브루트포스 방어 로직 부재")

# ===== 3. 로그아웃/토큰관리(브라우저) =====
add("로그아웃/토큰관리", "로그아웃 API 호출",
    "로그인 상태",
    "1) POST /api/users/logout 호출",
    "200 OK 반환(서버측 별도 무효화 로직 없음, TODO 상태임을 확인)",
    LOW, "AuthApiController.logout()은 TODO 상태, 서버 로직 없음")

add("로그아웃/토큰관리", "로그인 성공 시 localStorage 토큰 저장 확인",
    "브라우저에서 로그인 페이지 접근 가능",
    "1) login.html에서 정상 로그인\n2) 브라우저 DevTools > Application > LocalStorage 확인",
    "'token' 키에 JWT 값이 정상 저장됨",
    HIGH)

add("로그아웃/토큰관리", "세션 만료/로그아웃 시 localStorage 토큰 삭제 여부 (clearToken 버그 재현)",
    "로그인 상태로 대시보드 진입, localStorage에 'token' 키 존재",
    "1) 프로필 페이지에서 회원 탈퇴 실행 (또는 401 응답을 유도)\n2) clearToken() 호출 후 DevTools > LocalStorage에서 'token' 키 상태 확인",
    "'token' 키가 삭제되어야 함. 그러나 auth.js의 clearToken()이 localStorage.removeItem(\"toekn\")(오타)로 구현되어 있어 실제 'token' 키가 삭제되지 않을 것으로 예상됨 → 결함 재현 및 등록 대상",
    CRIT, "STEP1 핵심 예측 결함: auth.js clearToken() 오타(toekn≠token)")

add("로그아웃/토큰관리", "세션 만료(401) 발생 시 자동 로그아웃 및 리다이렉트",
    "로그인 상태에서 토큰을 만료/무효화시킬 수 있는 상태",
    "1) 토큰을 만료시키거나 서버측에서 무효화\n2) 보호된 화면에서 API 요청 발생시킴",
    "api.js가 401 응답을 감지하여 알림 표시 후 login 페이지로 이동. 단, clearToken() 결함으로 인해 실제 토큰은 삭제되지 않을 가능성 확인",
    HIGH, "TC-021과 연계 확인")

add("로그아웃/토큰관리", "브라우저 새로고침 후 로그인 상태 유지",
    "로그인 상태",
    "1) 대시보드에서 새로고침(F5) 실행",
    "localStorage의 토큰이 유지되어 로그인 상태(대시보드 화면)가 유지됨",
    MED)

add("로그아웃/토큰관리", "회원 탈퇴 후 잔존 토큰으로 API 재호출 시 응답 확인 (UserNotFoundException 500 매핑 재현)",
    "회원 탈퇴 직전 발급받은 유효 토큰 보유(만료 전)",
    "1) 프로필 페이지에서 회원 탈퇴 실행\n2) clearToken() 결함으로 인해 localStorage에 남아있는 토큰을 이용해 다시 GET /api/categories 등 보호된 API를 수동 호출(예: DevTools Console에서 fetch 재실행)",
    "정상적이라면 401/404가 반환되어야 하나, UserNotFoundException이 500 Internal Server Error로 매핑되어 있어 500 오류가 발생할 것으로 예상 → 결함 재현 및 등록 대상",
    CRIT, "STEP1 핵심 예측 결함: TC-021과 결합 시나리오, GlobalExceptionHandler UserNotFoundException→500")

# ===== 4. 회원 프로필 =====
add("회원 프로필", "사용자명 변경 성공",
    "로그인 상태",
    "1) PUT /api/users/me/username 호출, 유효한 userName 전달",
    "200 OK, 변경된 userName 응답에 반영",
    MED)

add("회원 프로필", "사용자명 공백으로 변경 시도",
    "로그인 상태",
    "1) userName=''으로 변경 요청",
    "400 Bad Request, '사용자명은 필수입니다' 메시지 반환",
    LOW)

add("회원 프로필", "비밀번호 변경 성공",
    "로그인 상태, 현재 비밀번호 알고 있음",
    "1) PUT /api/users/me/password 호출(현재 비밀번호 정확히 입력, 새 비밀번호 정책 충족)\n2) 새 비밀번호로 재로그인 시도",
    "200 OK 반환, 새 비밀번호로 로그인 성공",
    HIGH)

add("회원 프로필", "현재 비밀번호 불일치로 변경 실패",
    "로그인 상태",
    "1) currentPassword를 틀리게 입력하여 비밀번호 변경 요청",
    "400 Bad Request, '비밀번호가 일치하지 않습니다' 메시지 반환",
    MED)

add("회원 프로필", "새 비밀번호 정책 위반(8자 미만)",
    "로그인 상태",
    "1) newPassword를 7자 이하로 입력하여 변경 요청",
    "400 Bad Request, 비밀번호 길이 검증 메시지 반환",
    MED)

add("회원 프로필", "회원 탈퇴 성공",
    "로그인 상태, 비밀번호 알고 있음",
    "1) DELETE /api/users/me 호출(정확한 비밀번호 입력)\n2) 동일 계정으로 재로그인 시도",
    "204 No Content 반환, 이후 재로그인 시 401 Unauthorized",
    CRIT)

add("회원 프로필", "회원 탈퇴 시 비밀번호 불일치",
    "로그인 상태",
    "1) 잘못된 비밀번호로 탈퇴 요청",
    "400 Bad Request, '비밀번호가 일치하지 않습니다' 메시지 반환, 계정 유지",
    HIGH)

add("회원 프로필", "회원 탈퇴 시 연관 Category/Expense cascade 삭제 확인",
    "탈퇴 대상 계정에 카테고리 1개 이상, 지출 1건 이상 존재",
    "1) 카테고리/지출 데이터 생성\n2) 회원 탈퇴 실행\n3) DB 직접 조회(MySQL) 또는 관리자 관점에서 해당 user_id의 category/expense 레코드 잔존 여부 확인",
    "User 엔티티의 cascade=CascadeType.REMOVE 설정에 따라 연관된 Category/Expense 레코드가 모두 함께 삭제되어야 함(고아 레코드 없음)",
    CRIT, "STEP1 Gap: 기존 테스트는 재로그인 실패만 확인, DB 레벨 cascade 미검증")

# ===== 5. 카테고리 관리 =====
add("카테고리 관리", "카테고리 생성 성공",
    "로그인 상태",
    "1) POST /api/categories 호출, name='식비' 전달",
    "201 Created, 생성된 카테고리 정보(id, name) 반환",
    HIGH)

add("카테고리 관리", "카테고리 생성 시 이름 중복",
    "로그인 상태, 동일 사용자 소유의 '식비' 카테고리 이미 존재",
    "1) 동일 이름 '식비'로 카테고리 재생성 시도",
    "409 Conflict, '이미 존재하는 카테고리입니다' 메시지 반환",
    HIGH, "STEP1 Gap: 기존 테스트는 '수정' 중복만 커버, '생성' 중복 미검증")

add("카테고리 관리", "카테고리 생성 시 이름 공백",
    "로그인 상태",
    "1) name=''으로 카테고리 생성 요청",
    "400 Bad Request, '카테고리명은 필수입니다' 메시지 반환",
    MED)

add("카테고리 관리", "카테고리 목록 조회(본인 소유만)",
    "사용자 A, B 각각 카테고리 보유",
    "1) 사용자 A 토큰으로 GET /api/categories 호출",
    "사용자 A가 생성한 카테고리 목록만 반환되고 사용자 B의 카테고리는 포함되지 않음",
    HIGH)

add("카테고리 관리", "카테고리 수정 성공",
    "로그인 상태, 카테고리 1개 이상 보유",
    "1) PUT /api/categories/{id} 호출, 새 이름 전달",
    "200 OK, 변경된 이름 반영",
    MED)

add("카테고리 관리", "자기 자신과 동일한 이름으로 수정",
    "카테고리 이름 '식비' 보유",
    "1) 동일 카테고리를 동일 이름('식비')으로 수정 요청",
    "200 OK 반환(중복 예외 발생하지 않음)",
    LOW)

add("카테고리 관리", "타 카테고리와 이름 중복 시 수정 실패",
    "동일 사용자가 '식비', '교통비' 카테고리 보유",
    "1) '교통비' 카테고리를 '식비'로 수정 시도",
    "409 Conflict, '이미 존재하는 카테고리입니다' 메시지 반환",
    MED)

add("카테고리 관리", "타 사용자 카테고리 수정 시도",
    "사용자 A 소유 카테고리 존재, 사용자 B 로그인 토큰 보유",
    "1) 사용자 B 토큰으로 사용자 A의 categoryId를 대상으로 PUT 요청",
    "403 Forbidden, '해당 카테고리에 접근할 권한이 없습니다' 메시지 반환",
    CRIT)

add("카테고리 관리", "카테고리 삭제 성공(소속 지출 없음)",
    "지출이 연결되지 않은 카테고리 보유",
    "1) DELETE /api/categories/{id} 호출",
    "204 No Content 반환, 목록 조회 시 해당 카테고리 미노출",
    HIGH)

add("카테고리 관리", "카테고리 삭제 시 소속 지출 존재하면 삭제 거부",
    "특정 카테고리에 지출 1건 이상 등록된 상태",
    "1) 해당 카테고리에 지출을 1건 이상 생성\n2) 해당 카테고리 삭제 요청",
    "409 Conflict, '이 카테고리에 등록된 지출이 있어 삭제할 수 없습니다...' 메시지 반환, 카테고리/지출 데이터 유지",
    HIGH, "STEP1 핵심 Gap: 기존 자동화 테스트에 전혀 커버되지 않음(회귀 위험 최고)")

add("카테고리 관리", "타 사용자 카테고리 삭제 시도",
    "사용자 A 소유 카테고리 존재, 사용자 B 로그인 토큰 보유",
    "1) 사용자 B 토큰으로 사용자 A의 categoryId 삭제 요청",
    "403 Forbidden 반환, 카테고리 데이터 유지",
    CRIT)

add("카테고리 관리", "존재하지 않는 카테고리 접근",
    "로그인 상태",
    "1) 존재하지 않는 categoryId(예: 999999)로 조회/수정/삭제 요청",
    "404 Not Found, '카테고리를 찾을 수 없습니다' 메시지 반환",
    MED)

# ===== 6. 지출 관리 =====
add("지출 관리", "지출 생성 성공",
    "로그인 상태, 본인 소유 카테고리 존재",
    "1) POST /api/expenses 호출, categoryId/amount/content 정상값 전달",
    "201 Created, 생성된 지출 정보 반환(categoryName 포함)",
    CRIT)

add("지출 관리", "지출 생성 시 금액 0/음수 입력",
    "로그인 상태, 카테고리 존재",
    "1) amount=0 또는 amount=-1000으로 지출 생성 요청",
    "400 Bad Request, '금액은 0보다 커야 합니다' 메시지 반환",
    HIGH, "STEP1 Gap: 입력값 검증 실패 케이스 자동화 테스트 없음")

add("지출 관리", "지출 생성 시 금액 null",
    "로그인 상태, 카테고리 존재",
    "1) amount 필드 누락하여 지출 생성 요청",
    "400 Bad Request, '금액은 필수입니다' 메시지 반환",
    MED)

add("지출 관리", "지출 생성 시 content 공백",
    "로그인 상태, 카테고리 존재",
    "1) content=''으로 지출 생성 요청",
    "400 Bad Request, '내용은 필수입니다' 메시지 반환",
    MED)

add("지출 관리", "지출 생성 시 categoryId null",
    "로그인 상태",
    "1) categoryId 필드 누락하여 지출 생성 요청",
    "400 Bad Request, '카테고리는 필수입니다' 메시지 반환",
    MED)

add("지출 관리", "지출 생성 시 타 사용자 소유 categoryId 사용",
    "사용자 A 소유 카테고리 존재, 사용자 B 로그인 토큰 보유",
    "1) 사용자 B 토큰으로 지출 생성 시 categoryId에 사용자 A의 카테고리 id 사용",
    "403 Forbidden, '해당 카테고리에 접근할 권한이 없습니다' 메시지 반환(지출 생성 차단)",
    CRIT, "STEP1 Gap: 로직상 존재하나 자동화 테스트 미검증")

add("지출 관리", "존재하지 않는 categoryId로 지출 생성",
    "로그인 상태",
    "1) categoryId=999999(존재하지 않음)로 지출 생성 요청",
    "404 Not Found, '카테고리를 찾을 수 없습니다' 메시지 반환",
    MED)

add("지출 관리", "지출 목록 조회(본인 소유만)",
    "사용자 A, B 각각 지출 데이터 보유",
    "1) 사용자 A 토큰으로 GET /api/expenses 호출",
    "사용자 A의 지출만 반환되고 사용자 B의 지출은 포함되지 않음",
    HIGH)

add("지출 관리", "지출 수정 성공",
    "본인 소유 지출 존재",
    "1) PUT /api/expenses/{id} 호출, amount/content 등 변경",
    "200 OK, 변경된 값 반영",
    HIGH)

add("지출 관리", "지출 수정 시 카테고리 변경",
    "본인 소유 지출 및 2개 이상의 카테고리 보유",
    "1) 지출의 categoryId를 다른 본인 소유 카테고리로 변경",
    "200 OK, categoryName이 변경된 카테고리명으로 반영",
    MED)

add("지출 관리", "타 사용자 지출 수정 시도",
    "사용자 A 소유 지출 존재, 사용자 B 로그인 토큰 보유",
    "1) 사용자 B 토큰으로 사용자 A의 expenseId 수정 요청",
    "403 Forbidden, '해당 지출 내역에 접근할 권한이 없습니다' 메시지 반환",
    CRIT)

add("지출 관리", "지출 삭제 성공",
    "본인 소유 지출 존재",
    "1) DELETE /api/expenses/{id} 호출",
    "204 No Content 반환, 목록 조회 시 미노출",
    HIGH)

add("지출 관리", "타 사용자 지출 삭제 시도",
    "사용자 A 소유 지출 존재, 사용자 B 로그인 토큰 보유",
    "1) 사용자 B 토큰으로 사용자 A의 expenseId 삭제 요청",
    "403 Forbidden 반환, 지출 데이터 유지",
    CRIT)

add("지출 관리", "존재하지 않는 지출 접근",
    "로그인 상태",
    "1) 존재하지 않는 expenseId로 조회/수정/삭제 요청",
    "404 Not Found, '지출 내역을 찾을 수 없습니다' 메시지 반환",
    MED)

add("지출 관리", "지출 목록 categoryId 필터 동작",
    "서로 다른 카테고리에 지출 2건 이상 등록",
    "1) GET /api/expenses?categoryId={id} 호출",
    "해당 카테고리에 속한 지출만 반환됨",
    MED)

add("지출 관리", "지출 목록 날짜(startDate/endDate) 필터 동작",
    "여러 날짜에 걸쳐 지출 데이터 존재",
    "1) GET /api/expenses?startDate=YYYY-MM-DD&endDate=YYYY-MM-DD 호출",
    "지정 기간 내 생성된 지출만 반환됨(경계값 포함 여부 확인)",
    MED)

# ===== 7. 통계/요약 =====
add("통계/요약", "지출 요약 조회 성공",
    "로그인 상태, 여러 카테고리에 지출 데이터 존재",
    "1) GET /api/expenses/summary?startDate=...&endDate=... 호출",
    "200 OK, totalAmount(전체 합계)와 categorySummaryList(카테고리별 합계)가 실제 데이터와 일치",
    HIGH)

add("통계/요약", "지출 없는 기간 조회 시 총액 0 반환",
    "로그인 상태, 조회 기간 내 지출 데이터 없음",
    "1) 지출이 전혀 없는 기간으로 summary 조회",
    "totalAmount=0, categorySummaryList=[] 반환(500 오류 발생하지 않아야 함)",
    MED)

add("통계/요약", "startDate/endDate 파라미터 누락 시 처리",
    "로그인 상태",
    "1) startDate 또는 endDate 없이 GET /api/expenses/summary 호출",
    "@RequestParam 필수 파라미터 누락으로 400 Bad Request 반환 예상(실제 응답 및 에러 메시지 확인)",
    MED, "STEP1 Gap: 필수 파라미터 누락 케이스 미검증")

add("통계/요약", "startDate > endDate(날짜 역전) 케이스",
    "로그인 상태",
    "1) startDate가 endDate보다 이후 날짜인 조합으로 summary 조회",
    "서버가 빈 결과(0/empty)를 반환하는지, 400 오류로 처리하는지 실제 동작 확인 및 UX 관점에서 적절성 검토",
    MED, "STEP1 Gap: 날짜 역전 처리 로직 미검증")

add("통계/요약", "전월 대비 증감률 기능 확인",
    "README 상 '전월 대비 증감률' 기능 명시됨",
    "1) 대시보드 화면 및 GET /api/expenses/summary 응답에서 전월 대비 증감률 관련 필드/UI 존재 여부 확인",
    "ExpenseSummaryResponseDto/ExpenseService 코드상 관련 로직이 확인되지 않음 → 실제 미구현 여부 확인 후 기획 대비 누락 항목으로 별도 기록(버그 아님, 미구현 기능)",
    MED, "STEP1 기획-구현 불일치 후보")

add("통계/요약", "타 사용자 지출이 요약에 혼입되지 않는지 확인",
    "사용자 A, B 각각 다른 금액의 지출 데이터 보유",
    "1) 사용자 A 토큰으로 summary 조회\n2) totalAmount/categorySummaryList가 사용자 A의 데이터만 집계한 값인지 확인",
    "사용자 B의 지출 금액이 합산되지 않음(데이터 격리 확인)",
    CRIT)

# ===== 8. 프론트엔드(E2E) =====
add("프론트엔드(E2E)", "대시보드 도넛차트 렌더링 확인",
    "로그인 상태, 2개 이상 카테고리에 지출 데이터 존재",
    "1) 대시보드 페이지 진입\n2) 카테고리별 지출 비중 도넛차트 확인",
    "Chart.js 도넛차트가 카테고리별 비율에 맞게 정상 렌더링되고 범례/색상이 구분됨",
    MED)

add("프론트엔드(E2E)", "필터 폼(기간/카테고리) 제출 시 화면 갱신 확인",
    "로그인 상태, 다양한 기간/카테고리에 지출 데이터 존재",
    "1) 대시보드 필터 폼에서 기간 및 카테고리 선택 후 제출",
    "지출 목록, 총 소비 금액, 카테고리별 목록, 도넛차트가 필터 조건에 맞게 모두 갱신됨",
    HIGH)

add("프론트엔드(E2E)", "지출 등록 모달의 카테고리 select 동기화",
    "로그인 상태, 카테고리 1개 이상 보유",
    "1) '지출 추가' 모달 오픈\n2) 카테고리 select 옵션 확인",
    "현재 보유한 카테고리 목록이 select에 정상 노출됨",
    MED)

add("프론트엔드(E2E)", "지출 등록 시 categoryId 데이터 타입 이슈 확인",
    "로그인 상태, 카테고리 보유",
    "1) 대시보드 UI에서 지출 등록 폼 제출(select value가 문자열로 전송됨)\n2) 요청/응답 Network 탭 확인",
    "categoryId가 문자열로 전송되어도 백엔드에서 정상적으로 Long 타입으로 역직렬화되어 201 성공 응답을 받아야 함(타입 불일치로 인한 400 오류 발생 여부 확인)",
    HIGH, "STEP1 Gap: JSON.stringify({categoryId: select.value,...}) 문자열 전송 케이스")

add("프론트엔드(E2E)", "지출 삭제 confirm 및 목록 갱신 확인",
    "로그인 상태, 지출 데이터 존재",
    "1) 지출 목록에서 삭제 버튼 클릭\n2) confirm 다이얼로그에서 확인",
    "confirm 취소 시 삭제되지 않고, 확인 시 해당 지출이 삭제되며 목록/총액/차트가 갱신됨",
    LOW)

add("프론트엔드(E2E)", "카테고리 관리 목록에서 수정/삭제 버튼 동작",
    "로그인 상태, 카테고리 보유",
    "1) 카테고리 관리 목록에서 수정 버튼 클릭 후 이름 변경\n2) 삭제 버튼 클릭",
    "수정 시 모달에 기존 이름이 채워지고 저장 시 반영됨. 삭제 시 지출이 연결된 카테고리는 에러 메시지가 표시되고, 연결 안 된 카테고리는 정상 삭제됨",
    MED)

add("프론트엔드(E2E)", "미로그인 상태에서 보호된 페이지 접근 시 리다이렉트",
    "로그아웃 상태(또는 브라우저 시크릿 모드)",
    "1) 로그인하지 않은 상태로 /dashboard 또는 /profile 직접 접근",
    "isLoggedIn() 체크에 의해 login 페이지로 리다이렉트됨",
    MED)

add("프론트엔드(E2E)", "Swagger UI 인증 없이 접근 가능 여부 확인",
    "-",
    "1) 인증 없이 /swagger-ui.html 및 /v3/api-docs 접근",
    "현재 SecurityConfig 설정상 인증 없이 접근 가능함을 확인(운영 배포 시 노출 리스크로 별도 기록)",
    LOW, "STEP1 리스크: 운영 환경에서 Swagger 노출 여부 확인 필요")

add("프론트엔드(E2E)", "CORS 설정에 따른 프론트-백엔드 통신 확인",
    "프론트엔드가 localhost:8080 또는 127.0.0.1:5500에서 서빙",
    "1) 등록된 Origin에서 정상적으로 API 호출\n2) 등록되지 않은 임의 Origin(예: 다른 포트)에서 호출 시도(가능한 경우)",
    "등록된 Origin에서는 정상 동작, 미등록 Origin에서는 CORS 정책에 의해 차단됨을 확인",
    LOW)

add("프론트엔드(E2E)", "폼 유효성 오류 메시지 UI 노출 확인",
    "로그인 상태",
    "1) 지출/카테고리 등록 폼에서 서버측 검증 실패를 유발하는 값 입력 후 제출",
    "화면 내 에러 메시지 영역(errorMessage)에 서버에서 반환한 메시지가 사용자에게 노출됨",
    LOW)

# ---------------------------------------------------------------------------
# TC-ID 부여
# ---------------------------------------------------------------------------
for idx, row in enumerate(tcs, start=1):
    row[0] = f"TC-{idx:03d}"

total = len(tcs)
priority_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
for row in tcs:
    priority_counts[row[6]] += 1

print(f"총 TC 개수: {total}")
print(priority_counts)

# ---------------------------------------------------------------------------
# 엑셀 생성
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

priority_fill = {
    "Critical": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "High": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Medium": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Low": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
}

# ===== TC목록 시트 =====
ws = wb.active
ws.title = "TC목록"

columns = ["TC-ID", "카테고리", "테스트 항목", "전제 조건", "테스트 절차", "기대 결과", "우선순위", "결과", "비고"]
ws.append(columns)
for c in range(1, len(columns) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for row in tcs:
    ws.append(row)

# 서식 적용
widths = [10, 16, 30, 26, 42, 42, 11, 10, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in range(2, total + 2):
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        if c in (5, 6, 4, 9):
            cell.alignment = wrap
        elif c in (1, 2, 7, 8):
            cell.alignment = center
        else:
            cell.alignment = wrap
    # 우선순위 컬럼 색상
    priority_val = ws.cell(row=r, column=7).value
    fill = priority_fill.get(priority_val)
    if fill:
        ws.cell(row=r, column=7).fill = fill

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{total + 1}"

# ===== 요약 시트 =====
ws2 = wb.create_sheet("요약")

ws2["A1"] = "우선순위별 TC 현황"
ws2["A1"].font = Font(bold=True, size=12)

headers1 = ["우선순위", "전체 TC", "Pass", "Fail", "Blocked"]
ws2.append([])
ws2.append(headers1)
header_row_idx = 3
for c in range(1, len(headers1) + 1):
    cell = ws2.cell(row=header_row_idx, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

priority_order = ["Critical", "High", "Medium", "Low"]
r = header_row_idx + 1
for p in priority_order:
    cnt = priority_counts[p]
    ws2.cell(row=r, column=1, value=p)
    ws2.cell(row=r, column=2, value=cnt)
    ws2.cell(row=r, column=3, value=0)
    ws2.cell(row=r, column=4, value=0)
    ws2.cell(row=r, column=5, value=0)
    for c in range(1, 6):
        ws2.cell(row=r, column=c).border = border
        ws2.cell(row=r, column=c).alignment = center
    r += 1

total_row = r
ws2.cell(row=total_row, column=1, value="합계")
ws2.cell(row=total_row, column=2, value=total)
ws2.cell(row=total_row, column=3, value=0)
ws2.cell(row=total_row, column=4, value=0)
ws2.cell(row=total_row, column=5, value=0)
for c in range(1, 6):
    cell = ws2.cell(row=total_row, column=c)
    cell.font = Font(bold=True)
    cell.border = border
    cell.alignment = center

# 전체 진행 현황
start2 = total_row + 3
ws2.cell(row=start2, column=1, value="전체 진행 현황").font = Font(bold=True, size=12)

headers2 = ["항목", "값"]
r2 = start2 + 1
ws2.append([]) if False else None
for c, h in enumerate(headers2, start=1):
    cell = ws2.cell(row=r2, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

progress_rows = [
    ("전체 TC수", total),
    ("Pass", 0),
    ("Fail(미해결)", 0),
    ("Blocked", 0),
    ("미실행", total),
    ("통과율", "0%"),
]

rr = r2 + 1
for label, value in progress_rows:
    ws2.cell(row=rr, column=1, value=label)
    ws2.cell(row=rr, column=2, value=value)
    for c in range(1, 3):
        ws2.cell(row=rr, column=c).border = border
        ws2.cell(row=rr, column=c).alignment = center if c == 2 else Alignment(horizontal="left", vertical="center")
    rr += 1

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12

wb.save(r"C:\Users\jyr\Desktop\study\dev\portfolio\MoneyLog\QA\테스트케이스.xlsx")
print("저장 완료: 테스트케이스.xlsx")
