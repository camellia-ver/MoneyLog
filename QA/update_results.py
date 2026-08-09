# -*- coding: utf-8 -*-
"""STEP3: 테스트케이스.xlsx 결과 컬럼 갱신 + 요약 시트 갱신"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

PATH = r"C:\Users\jyr\Desktop\study\dev\portfolio\MoneyLog\QA\테스트케이스.xlsx"

# TC-ID -> (결과, 비고 추가문구 또는 None(기존 비고 유지))
results = {}

def setr(tcid, result, note=None):
    results[tcid] = (result, note)

PASS = "Pass"
FAIL = "Fail"
BLOCK = "Blocked"

# ---- 회원가입 ----
setr("TC-001", PASS)
setr("TC-002", PASS)
setr("TC-003", PASS)
setr("TC-004", PASS)
setr("TC-005", PASS)
setr("TC-006", PASS)
setr("TC-007", PASS)
setr("TC-008", PASS, "대문자 이메일도 동일 계정으로 간주되어 409 반환됨(대소문자 무관 중복 차단, 정상 동작으로 판단)")

# ---- 로그인/인증 ----
setr("TC-009", PASS)
setr("TC-010", PASS)
setr("TC-011", PASS)
setr("TC-012", PASS, "DTO 검증 어노테이션은 없지만 서비스 로직에서 InvalidCredentialsException으로 안전하게 401 처리됨")
setr("TC-013", PASS, "위와 동일, 공백 비밀번호도 401로 안전하게 처리됨")
setr("TC-014", PASS)
setr("TC-015", PASS)
setr("TC-016", PASS)
setr("TC-017", BLOCK, "커버리지 한계: jwt.expiration=1시간이라 실시간 대기 불가. TC-016(위조 토큰)과 동일한 JwtException catch 경로를 타는 것을 코드로 확인했으나 실측은 미수행")
setr("TC-018", PASS, "[리스크 발견] 5회 연속 로그인 실패 후에도 계정 잠금/지연 없이 정상 로그인 가능 - 브루트포스 방어 부재(결함보고서 Medium 항목 등록)")

# ---- 로그아웃/토큰관리 ----
setr("TC-019", PASS)
setr("TC-020", PASS, "headless Chrome에서 실제 auth.js(saveToken)를 로드해 localStorage 저장 확인")
setr("TC-021", FAIL, "[Critical 결함] headless Chrome에서 실제 auth.js를 로드해 clearToken() 실행 -> localStorage.removeItem(\"toekn\")(오타)로 인해 'token' 키가 삭제되지 않음을 실제로 재현함. 결함보고서 DEF-001 참조")
setr("TC-022", FAIL, "[Critical 결함 연계] headless Chrome에서 실제 api.js+실제 백엔드(위조토큰->401)를 이용해 재현: apiRequest()의 401 처리 로직은 정상 호출되나 내부 clearToken()이 실패하여 토큰이 남음. 결함보고서 DEF-001 참조")
setr("TC-023", PASS, "localStorage는 브라우저 표준 사양상 새로고침에도 유지됨(코드상 getToken()이 매번 localStorage를 읽음) - 코드 검토로 판정")
setr("TC-024", FAIL, "[Critical 결함] 실제 API로 재현: 회원탈퇴 성공(204) 직후 잔존 토큰으로 GET /api/categories 호출 시 500 Internal Server Error 반환(UserNotFoundException->500 매핑). 결함보고서 DEF-002 참조")

# ---- 회원 프로필 ----
setr("TC-025", PASS)
setr("TC-026", PASS)
setr("TC-027", PASS)
setr("TC-028", PASS)
setr("TC-029", PASS)
setr("TC-030", PASS, "카테고리/지출이 없는 계정(userC)으로 실제 탈퇴 성공(204) 및 재로그인 401 확인")
setr("TC-031", PASS)
setr("TC-032", FAIL, "[Critical 신규 결함] 카테고리+지출이 모두 있는 계정(userD)으로 실제 탈퇴 시도 시 DataIntegrityViolationException 발생, 계정 삭제 자체가 실패함(트랜잭션 롤백으로 데이터 유실은 없으나 탈퇴 기능이 동작하지 않음). 결함보고서 DEF-003 참조(최우선)")

# ---- 카테고리 관리 ----
setr("TC-033", PASS)
setr("TC-034", PASS, "STEP1 Gap이었던 생성 시 중복 케이스 실제 검증 완료, 409 정상 반환")
setr("TC-035", PASS)
setr("TC-036", PASS)
setr("TC-037", PASS)
setr("TC-038", PASS)
setr("TC-039", PASS)
setr("TC-040", PASS)
setr("TC-041", PASS)
setr("TC-042", PASS, "STEP1 Gap이었던 케이스 실제 검증 완료, 409 정상 반환(단, 회원탈퇴 cascade 경로에서는 동일 보호가 우회되어 DEF-003 발생 - 별개 코드 경로)")
setr("TC-043", PASS)
setr("TC-044", PASS)

# ---- 지출 관리 ----
setr("TC-045", PASS)
setr("TC-046", PASS)
setr("TC-047", PASS)
setr("TC-048", PASS)
setr("TC-049", PASS)
setr("TC-050", PASS, "STEP1 Gap이었던 케이스 실제 검증 완료, 403 정상 반환")
setr("TC-051", PASS)
setr("TC-052", PASS)
setr("TC-053", PASS)
setr("TC-054", PASS)
setr("TC-055", PASS)
setr("TC-056", PASS)
setr("TC-057", PASS)
setr("TC-058", PASS)
setr("TC-059", PASS)
setr("TC-060", PASS)

# ---- 통계/요약 ----
setr("TC-061", PASS)
setr("TC-062", PASS)
setr("TC-063", FAIL, "[Critical 신규 결함] startDate/endDate 파라미터 누락 시 MissingServletRequestParameterException이 발생하나 클라이언트에는 400이 아닌 401(빈 바디)이 반환됨. 결함보고서 DEF-004 참조(시스템 전역 영향)")
setr("TC-064", PASS, "startDate>endDate 역전 시 에러 없이 빈 결과(totalAmount=0) 반환. 기능은 깨지지 않으나 사용자 피드백 부재는 UX 개선 여지로 기록")
setr("TC-065", PASS, "실제 응답에 전월 대비 증감률 관련 필드가 없음을 확인함 - 버그 아닌 기획 대비 미구현 항목으로 결함보고서에 정보성 기록")
setr("TC-066", PASS)

# ---- 프론트엔드(E2E) ----
setr("TC-067", BLOCK, "커버리지 한계: Chart.js 도넛차트의 시각적 렌더링(canvas) 검증은 브라우저 자동화(Selenium/Playwright) 도구 부재로 수행 불가. dashboard.js 코드 정적 리뷰로 로직만 확인함")
setr("TC-068", BLOCK, "커버리지 한계: 필터 폼 제출->DOM 갱신의 UI 상호작용 검증 불가. 단, 백엔드 필터링 자체(TC-059,060)는 API 레벨에서 정상 확인됨")
setr("TC-069", BLOCK, "커버리지 한계: 모달 오픈 시 select 옵션 동기화는 DOM 상호작용 검증 도구 부재로 수행 불가")
setr("TC-070", PASS, "실제 API에 categoryId를 문자열로 전송(\"39\")해도 201 성공 확인 - Jackson이 문자열->Long 정상 변환함(STEP1 우려 Gap 해소, 결함 아님)")
setr("TC-071", BLOCK, "커버리지 한계: confirm() 다이얼로그 및 목록 갱신 UI 상호작용 검증 도구 부재")
setr("TC-072", BLOCK, "커버리지 한계: 카테고리 관리 목록 버튼 클릭 상호작용 검증 도구 부재")
setr("TC-073", PASS, "headless Chrome에서 실제 auth.js의 isLoggedIn()이 미로그인 상태에서 false를 반환함을 확인, dashboard.js/profile.js의 리다이렉트 가드 코드 경로와 결합하여 판정")
setr("TC-074", PASS, "실제 확인: 인증 없이 /swagger-ui/index.html, /v3/api-docs 모두 200 응답. [리스크 발견] 운영 배포 시 API 문서 노출 위험 - 결함보고서 Low 항목 등록")
setr("TC-075", PASS, "실제 확인: 허용 Origin(127.0.0.1:5500)은 CORS 통과(200), 미허용 Origin은 403 Invalid CORS request로 정상 차단됨")
setr("TC-076", BLOCK, "커버리지 한계: 폼 에러 메시지 DOM 노출의 시각적 확인은 브라우저 자동화 도구 부재로 수행 불가. 단, 백엔드 에러 메시지 자체(message 필드)는 각 API 테스트에서 확인됨")

wb = openpyxl.load_workbook(PATH)
ws = wb["TC목록"]

RESULT_COL = 8  # H
NOTE_COL = 9    # I

pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
block_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
fill_map = {PASS: pass_fill, FAIL: fail_fill, BLOCK: block_fill}

priority_counts = Counter()
result_by_priority = {}  # priority -> Counter(result)

missing = []
for r in range(2, ws.max_row + 1):
    tcid = ws.cell(row=r, column=1).value
    priority = ws.cell(row=r, column=7).value
    if tcid not in results:
        missing.append(tcid)
        continue
    result, note = results[tcid]
    ws.cell(row=r, column=RESULT_COL, value=result)
    ws.cell(row=r, column=RESULT_COL).fill = fill_map[result]
    ws.cell(row=r, column=RESULT_COL).alignment = Alignment(horizontal="center", vertical="center")
    if note:
        existing = ws.cell(row=r, column=NOTE_COL).value
        combined = note if not existing else f"{existing} | {note}"
        ws.cell(row=r, column=NOTE_COL, value=combined)
        ws.cell(row=r, column=NOTE_COL).alignment = Alignment(wrap_text=True, vertical="top")

    priority_counts[priority] += 1
    result_by_priority.setdefault(priority, Counter())[result] += 1

if missing:
    raise SystemExit(f"매핑 누락된 TC 존재: {missing}")

total = sum(priority_counts.values())
total_pass = sum(c[PASS] for c in result_by_priority.values())
total_fail = sum(c[FAIL] for c in result_by_priority.values())
total_block = sum(c[BLOCK] for c in result_by_priority.values())

print("TOTAL", total, "PASS", total_pass, "FAIL", total_fail, "BLOCKED", total_block)
for p in ["Critical", "High", "Medium", "Low"]:
    c = result_by_priority.get(p, Counter())
    print(p, dict(c))

# ===== 요약 시트 갱신 =====
ws2 = wb["요약"]

priority_order = ["Critical", "High", "Medium", "Low"]
header_row_idx = 3
r = header_row_idx + 1
for p in priority_order:
    c = result_by_priority.get(p, Counter())
    ws2.cell(row=r, column=1, value=p)
    ws2.cell(row=r, column=2, value=priority_counts.get(p, 0))
    ws2.cell(row=r, column=3, value=c[PASS])
    ws2.cell(row=r, column=4, value=c[FAIL])
    ws2.cell(row=r, column=5, value=c[BLOCK])
    r += 1

total_row = r
ws2.cell(row=total_row, column=1, value="합계")
ws2.cell(row=total_row, column=2, value=total)
ws2.cell(row=total_row, column=3, value=total_pass)
ws2.cell(row=total_row, column=4, value=total_fail)
ws2.cell(row=total_row, column=5, value=total_block)

# 전체 진행 현황 섹션 갱신
start2 = total_row + 3
r2 = start2 + 1
progress_rows = [
    ("전체 TC수", total),
    ("Pass", total_pass),
    ("Fail(미해결)", total_fail),
    ("Blocked", total_block),
    ("미실행", 0),
    ("통과율", f"{round(total_pass/total*100, 1)}%"),
]
rr = r2 + 1
for label, value in progress_rows:
    ws2.cell(row=rr, column=1, value=label)
    ws2.cell(row=rr, column=2, value=value)
    rr += 1

wb.save(PATH)
print("저장 완료")
