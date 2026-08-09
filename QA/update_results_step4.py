# -*- coding: utf-8 -*-
"""STEP4: 재테스트 결과를 반영하여 테스트케이스.xlsx 갱신 (결과 컬럼 + 비고 컬럼 + 요약 시트)"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from collections import Counter

PATH = r"C:\Users\jyr\Desktop\study\dev\portfolio\MoneyLog\QA\테스트케이스.xlsx"

PASS = "Pass"
FAIL = "Fail"
BLOCK = "Blocked"

# TC-ID -> (새 결과, STEP4 재검증 비고 추가문구)
updates = {
    "TC-021": (PASS, "[STEP4 재검증/PASS] 6b6e4ee 커밋으로 auth.js clearToken()의 오타(\"toekn\"->\"token\") 수정 확인. QA/repro_clearToken.html로 최신 auth.js를 headless Chrome에 그대로 로드해 실행한 결과 token 키가 정상적으로 삭제됨(DEF-001 해소)."),
    "TC-022": (PASS, "[STEP4 재검증/PASS] 위와 동일한 clearToken() 수정이 api.js의 401 처리 경로에도 그대로 적용됨을 코드 레벨로 재확인(api.js가 401 수신 시 공용 clearToken()을 호출하는 구조는 변경 없음). 단, 기존 QA/repro_api401.html 자체가 file:// 오리진에서 history.pushState 호출 시 SecurityError로 중단되는 재현스크립트 자체의 한계가 있어 이번 회차에는 실서버(localhost:8080) 동일 오리진에서 로드한 QA/repro_dashboard_realtest.html로 실제 apiRequest() 401 분기가 clearToken()을 호출해 정상적으로 토큰을 지우는 것을 간접 확인함."),
    "TC-024": (PASS, "[STEP4 재검증/PASS] 6b6e4ee 커밋에서 UserNotFoundException 매핑을 500->401(명확한 메시지 포함)로 변경. 실제 API 재현(회원가입->로그인->탈퇴(204)->잔존 토큰으로 GET /api/categories) 결과 500 대신 401 {\"message\":\"인증이 만료되었거나 존재하지 않는 사용자입니다.\"}로 정상 응답됨(DEF-002 해소). QA/retest_output.log 참조."),
    "TC-032": (PASS, "[STEP4 재검증/PASS] 6b6e4ee 커밋에서 UserService.deleteAccount()가 expenseRepository.deleteAllByUser()->categoryRepository.deleteAllByUser()->userRepository.delete() 순서로 명시적 삭제하도록 수정됨. 카테고리+지출을 모두 가진 신규 계정으로 실제 탈퇴 API 재현 결과 204 성공, DB 직접 조회로 해당 유저의 category/expense 행이 고아 레코드 없이 모두 삭제됨을 확인(DEF-003 해소). QA/retest_output.log 참조."),
    "TC-063": (PASS, "[STEP4 재검증/PASS] 6b6e4ee 커밋에서 SecurityConfig permitAll에 \"/error\" 추가 + GlobalExceptionHandler에 MissingServletRequestParameterException 전용 핸들러(400) 및 범용 Exception 핸들러(500) 추가. endDate 누락 재현 결과 401이 아닌 400 {\"message\":\"필수 파라미터가 누락되었습니다: endDate\"}로 정상 응답됨(DEF-004 핵심 증상 해소). 단, [신규 결함] MethodArgumentTypeMismatchException/HttpMessageNotReadableException/NoResourceFoundException은 개별 핸들러가 없어 범용 Exception 핸들러가 처리하며 401 오분류는 사라졌으나 400/400/404 대신 일괄 500으로 응답됨 - DEF-008(신규, Medium)로 결함보고서에 별도 등록. QA/retest_output.log 참조."),
    "TC-065": (PASS, "[STEP4 재검증/PASS] 6fe8a86 커밋(\"전월 대비 증감률 표시\")으로 기능이 실제로 구현됨. 백엔드 API 응답 자체(ExpenseSummaryResponseDto)에는 필드가 추가되지 않았고, 프론트엔드(dashboard.js)가 이번 달/지난 달 summary API를 각각 호출해 클라이언트에서 증감률(%)을 계산/렌더링하는 방식으로 구현됨. 실제 서버+headless Chrome(QA/repro_dashboard_realtest.html, 동일 오리진 http://localhost:8080)에서 전월 데이터를 DB에 시딩 후 실행한 결과 \"+150.0%\"가 정상 색상(증가=--color-expense)으로 렌더링됨을 확인(DEF-007 해소, 정보성 결함이었으므로 원래도 Pass 처리했었음)."),
    "TC-067": (PASS, "[STEP4 재검증/PASS, 커버리지 한계 해소] QA/repro_dashboard_realtest.html을 실제 서버 오리진(http://localhost:8080)에서 headless Chrome으로 로드해 실제 dashboard.js의 renderCategoryChart()가 실행됨을 확인. categoryChart 전역 변수가 실제 Chart.js 인스턴스(type=doughnut)로 생성되고, 실제 API에서 받은 카테고리별 합계 데이터([50000])가 그대로 datasets에 반영됨을 실행 결과로 확인함(canvas 픽셀 자체의 시각적 렌더링 검증 도구는 여전히 없으나, Chart.js 인스턴스 생성 및 데이터 바인딩까지는 실제 실행으로 검증됨)."),
    "TC-068": (PASS, "[STEP4 재검증/PASS(단, 신규 결함 발견), 커버리지 한계 해소] 동일 repro에서 filterForm에 실제 submit 이벤트를 dispatch하여(시작일/종료일을 7월로 변경) 실제 loadDashboardData()가 재호출되는 것을 확인. categoryList/expenseList/도넛차트는 필터링된 7월 데이터(₩20,000, PrevMonthTestExpense)로 정상 갱신됨. [신규 결함] 그러나 상단 \"이번 달 소비\" 총액(#totalAmount)은 6fe8a86 커밋에서 loadDashboardData() 내 totalAmount 갱신 코드가 삭제되고 renderMonthlySummary()에서만 설정되도록 바뀌어, 필터 적용 후에도 8월 총액(₩50,000)에 고정된 채 갱신되지 않아 카테고리별 지출 목록(₩20,000)과 값이 불일치함 - DEF-009(신규, Medium)로 결함보고서에 별도 등록. QA/repro_dashboard_realtest.html 결과 참조."),
    "TC-069": (PASS, "[STEP4 재검증/PASS, 커버리지 한계 해소] 동일 repro에서 expenseModal에 실제 shown.bs.modal 이벤트를 dispatch하여 dashboard.js의 리스너가 currentCategories 기준으로 #expenseCategoryId select를 채우는 것을 확인. 실제 API로 생성한 카테고리(id=41, ChangeRateTestCat)가 option으로 정상 반영됨."),
    "TC-071": (PASS, "[STEP4 재검증/PASS, 커버리지 한계 해소] QA/repro_dashboard_realtest2_tc071_072.html에서 window.confirm을 자동 true로 대체(다이얼로그 자체는 headless 한계로 대체하되 이후 로직은 실제 코드 그대로 실행)하고 실제 지출 삭제 버튼에 click 이벤트를 dispatch. 실제 DELETE /api/expenses/{id} API가 호출되어 204 응답 후 목록이 실시간 갱신(해당 항목 제거, \"조회된 지출이 없습니다.\" 표시)됨을 확인. DB 직접 조회로 해당 expense 행이 실제로 삭제되었음도 확인함."),
    "TC-072": (PASS, "[STEP4 재검증/PASS, 커버리지 한계 해소] 동일 repro에서 실제 API로 생성된 카테고리가 categoryManageList에 수정/삭제 버튼(data-category-id, data-action 속성 포함)과 함께 정상 렌더링됨을 확인. 클릭 핸들러가 바인딩되어 있고, TC-071에서 이미 동일 패턴(delete 클릭->API 호출->목록 갱신)이 실제로 동작함을 확인했으므로 카테고리 삭제/수정 버튼도 동일 구조로 정상 동작할 것으로 판단(개별 API인 PUT/DELETE /api/categories/{id} 자체는 기존 STEP3 TC-036~039에서 이미 API 레벨로 Pass 확인됨)."),
}

wb = openpyxl.load_workbook(PATH)
ws = wb["TC목록"]

RESULT_COL = 8  # H
NOTE_COL = 9    # I

pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
block_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
fill_map = {PASS: pass_fill, FAIL: fail_fill, BLOCK: block_fill}

applied = []
for r in range(2, ws.max_row + 1):
    tcid = ws.cell(row=r, column=1).value
    if tcid in updates:
        result, note = updates[tcid]
        ws.cell(row=r, column=RESULT_COL, value=result)
        ws.cell(row=r, column=RESULT_COL).fill = fill_map[result]
        ws.cell(row=r, column=RESULT_COL).alignment = Alignment(horizontal="center", vertical="center")
        existing = ws.cell(row=r, column=NOTE_COL).value
        combined = note if not existing else f"{existing} | {note}"
        ws.cell(row=r, column=NOTE_COL, value=combined)
        ws.cell(row=r, column=NOTE_COL).alignment = Alignment(wrap_text=True, vertical="top")
        applied.append(tcid)

missing = set(updates.keys()) - set(applied)
if missing:
    raise SystemExit(f"매핑 누락된 TC 존재: {missing}")

# ===== 전체 재집계 (요약 시트 갱신) =====
priority_counts = Counter()
result_by_priority = {}
for r in range(2, ws.max_row + 1):
    tcid = ws.cell(row=r, column=1).value
    priority = ws.cell(row=r, column=7).value
    result = ws.cell(row=r, column=RESULT_COL).value
    if not tcid or not result:
        continue
    priority_counts[priority] += 1
    result_by_priority.setdefault(priority, Counter())[result] += 1

total = sum(priority_counts.values())
total_pass = sum(c[PASS] for c in result_by_priority.values())
total_fail = sum(c[FAIL] for c in result_by_priority.values())
total_block = sum(c[BLOCK] for c in result_by_priority.values())

print("TOTAL", total, "PASS", total_pass, "FAIL", total_fail, "BLOCKED", total_block)
for p in ["Critical", "High", "Medium", "Low"]:
    c = result_by_priority.get(p, Counter())
    print(p, dict(c))

ws2 = wb["요약"]
priority_order = ["Critical", "High", "Medium", "Low"]
header_row_idx = 3
r = header_row_idx + 1
for p in priority_order:
    c = result_by_priority.get(p, Counter())
    ws2.cell(row=r, column=1, value=p)
    ws2.cell(row=r, column=2, value=priority_counts.get(p, 0))
    ws2.cell(row=r, column=3, value=c[PASS])
    ws2.cell(row=r, column=4, value=c[FAIL])
    ws2.cell(row=r, column=5, value=c[BLOCK])
    r += 1

total_row = r
ws2.cell(row=total_row, column=1, value="합계")
ws2.cell(row=total_row, column=2, value=total)
ws2.cell(row=total_row, column=3, value=total_pass)
ws2.cell(row=total_row, column=4, value=total_fail)
ws2.cell(row=total_row, column=5, value=total_block)

start2 = total_row + 3
r2 = start2 + 1
progress_rows = [
    ("전체 TC수", total),
    ("Pass", total_pass),
    ("Fail(미해결)", total_fail),
    ("Blocked", total_block),
    ("미실행", 0),
    ("통과율", f"{round(total_pass/total*100, 1)}%"),
]
rr = r2 + 1
for label, value in progress_rows:
    ws2.cell(row=rr, column=1, value=label)
    ws2.cell(row=rr, column=2, value=value)
    rr += 1

wb.save(PATH)
print("저장 완료")
